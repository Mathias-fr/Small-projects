import openpyxl

def get_data_from_excel_file(filename):
#Acquires two dimensional data from excel file and returns the data in a list.

  workbook = openpyxl.load_workbook(filename)
  sheet = workbook.active
  max_row = sheet.max_row
  max_column = sheet.max_column

  print(f"Max row is {max_row}")
  print(f"Max column is {max_column}")

  for i in range(1, max_row + 1):
    for j in range(1, max_column + 1):
      if sheet.cell(row=i, column=j).value != None:
        top_left_row = i
        top_left_column = j
        print(f"Top left cell is row {top_left_row}, column {top_left_column}")
        break
    else:
      continue
    break

  column_names = []
  for i in range(top_left_column, max_column + 1):
    column_names.append(sheet.cell(row=top_left_row, column=i).value)

  print(f"Column names are: {column_names}")
  print(f"Number of rows containing data is {max_row - top_left_row}")

  data = []
  for i in range(top_left_row + 1, max_row + 1):
    dict = {}
    if sheet.cell(row=i, column=top_left_column).value == None:
      continue
    for j in range(top_left_column, max_column + 1):
      dict[column_names[j - top_left_column]] = sheet.cell(row=i, column=j).value
    data.append(dict)

  print("Data has been collected")

  return data
