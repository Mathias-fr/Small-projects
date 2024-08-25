import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

def NVDB_road_reference_searcher():
#Acquires a list of road references from an excel file and marks the location of the
#road references in the NVDB website.
    
    workbook = openpyxl.load_workbook("Bru_data.xlsx")
    sheet = workbook.active
    max_row = sheet.max_row
    road_reference_column = 7
    
    road_references = []
    for i in range(1, max_row + 1):
        road_references.append(sheet.cell(row=i, column=road_reference_column).value)

    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)

    browser = webdriver.Chrome(options=options)
    browser.get("https://vegkart.atlas.vegvesen.no/#kartlag:geodata/@418362,7306793,9")
    browser.implicitly_wait(15)

    element = browser.find_element(By.ID, "close_changelog_header_button")
    element.click()
    browser.implicitly_wait(15)

    element = browser.find_element(By.ID, "close_splash_header")
    element.click()
    browser.implicitly_wait(15)

    element = browser.find_element(By.ID, "searchText")
    element.send_keys("Vefsn")
    element.send_keys(Keys.RETURN)
    browser.implicitly_wait(15)

    for i in road_references:
        element = browser.find_element(By.ID, "searchText")
        element.send_keys(i)
        element.send_keys(Keys.RETURN)
        browser.implicitly_wait(15)

    print("The road reference locations have been marked in NVDB.")
